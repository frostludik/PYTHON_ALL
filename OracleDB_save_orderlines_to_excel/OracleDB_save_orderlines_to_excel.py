"""
author: Ludek Mraz
email: ludek.mraz@centrum.cz
discord: Luděk M.#5570
"""

# This code runs SQL query against Oracle database of WMS
# Extracts data
# Saves them in .XLSX
#
# Added:
#   Check connection do destination network folder, stop program after 5 retries
#   Check if the file exists already
#   Saving progress to a logfile in "C:\\temp"
#   Function print_with_timestamp

import oracledb
import openpyxl
import os.path
import datetime
import os
import time
import sys
import secrets

query = """
    SELECT sysdate,lk250val
    FROM LK250T1
    WHERE lk250kpi = 6050
    """
file_path = r"\\10.13.245.1\dc\autosave\Orderlines.xlsx"
now = datetime.datetime.now()
time_str = now.strftime("%Y-%m-%d %H:%M:%S")
log_time_string = now.strftime("%Y%m%d_%H%M%S")
max_retries = 5
retry_delay = 15
retry_count = 0
log_file_path = os.path.join("C:\\temp", f'log_WMS_save_orderlines_{log_time_string}.txt')


def print_with_timestamp(msg, file=sys.stdout):
    '''
    adds timestamp to each print statement
    Example usage:
    print_with_timestamp("Hello, world!", file=log_file )
    '''
    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")
    print(f"[{timestamp}] {msg}", file=file)
    log_file.flush()


#open log file
log_file = open(log_file_path, "w")
sys.stdout = log_file

#try connection to destination folder
while True:
    retry_count += 1
    try:
        files = os.listdir(r"\\10.13.245.1\dc\Autosave")
        print_with_timestamp("File destination folder is reachable.", file=log_file)
        break        
    except OSError as e:
        print_with_timestamp(f"Error accessing network path on attempt {retry_count}/{max_retries}: {e} \
        \nTrying to connect again....", file=log_file)
        
        if retry_count == max_retries:
            print_with_timestamp("Failed to access network path. Closing program.", file=log_file)
            exit()
        else:
            print_with_timestamp(f"Retrying in {retry_delay} seconds...", file=log_file)
            time.sleep(retry_delay)
            
        
#open xlsx file
if os.path.isfile(file_path):
    workbook = openpyxl.load_workbook(file_path)
    print_with_timestamp("File already exists. Continue...", file=log_file)
else:
    workbook = openpyxl.Workbook()
    workbook.save(file_path)
    print_with_timestamp(f"File not exists \nFile successfully created in {file_path}", file=log_file)
    
    
#add column names
worksheet = workbook.active
column_names = ['system_date_time', 'Picked_orderlines']
for i in range(len(column_names)):
    worksheet.cell(row=1, column=i+1, value=column_names[i])

#set fixed columns' width
for col in ['A', 'B']:
    worksheet.column_dimensions[col].width = 18

#find the first empty row in worksheet
row_num = 1
for cell in worksheet["A"]:
    if cell.value is None:
        break
else:
    row_num = cell.row + 1

#connect to database and run query
db = oracledb.connect('USER/password@hostname:dc-wms01.dc.company.int/database05')
print_with_timestamp("Successfully connected to Oracle Database", file=log_file)  

cursor = db.cursor()
cursor.execute(query)

#write data to xlsx
for row in cursor:
    worksheet.cell(row=row_num, column=1).value = row[0]
    worksheet.cell(row=row_num, column=2).value = row[1]
    row_num += 1
    print_with_timestamp("Writing data to xlsx", file=log_file)

#close database
db.close()
print_with_timestamp("Database closed", file=log_file)

#save xlsx changes
workbook.save(file_path)
workbook.close()
print_with_timestamp("Xlsx file saved", file=log_file)

# Close the log file
log_file.close()